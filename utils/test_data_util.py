import openpyxl as xl
from utils.search_data_type import search_test_data


@staticmethod
def get_test_data_record(filename, row_index):
    wb = xl.load_workbook(filename)
    sheet = wb['Search']
    search_test_data["initial_search_value"] = sheet.cell(row_index, 2).value
    search_test_data["language"] = sheet.cell(row_index, 3).value
    search_test_data["stars"] = sheet.cell(row_index, 4).value
    search_test_data["license_index"] = int(sheet.cell(row_index, 5).value)
    search_test_data["followers"] = sheet.cell(row_index, 6).value
    search_test_data["expected_result"] = sheet.cell(row_index, 8).value
    return search_test_data

